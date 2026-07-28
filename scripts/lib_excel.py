"""Excel sheet writers for PSI Cash Flow Report."""
from __future__ import annotations

from calendar import monthrange
from collections import defaultdict
from datetime import datetime

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from lib_pivot import INDONESIAN_MONTHS, MONTHS_FULL, month_label


# Styling palette
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, color="1F3864")
SUBTOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
SUBTOTAL_FONT = Font(bold=True)
SECTION_TOTAL_FILL = PatternFill("solid", fgColor="FFF8E1")
SECTION_TOTAL_FONT = Font(bold=True, color="1F3864")
ZEBRA_FILL = PatternFill("solid", fgColor="FAFBFD")
COMPANY_FONT = Font(bold=True, size=14, color="1F3864")
TOTAL_FILL = PatternFill("solid", fgColor="FFE699")
TOTAL_FONT = Font(bold=True)

# Borders
_thin = Side(style="thin", color="BFC8DA")
_medium = Side(style="medium", color="1F3864")
BORDER_ALL = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_HEADER = Border(left=_thin, right=_thin, top=_medium, bottom=_medium)
BORDER_TOTAL_TOP = Border(left=_thin, right=_thin, top=_medium, bottom=_thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def _apply_style(cell, kind):
    cell.border = BORDER_ALL
    if kind == "section_header":
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    elif kind == "subsection":
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
    elif kind in ("subtotal_section",):
        cell.fill = SUBTOTAL_FILL
        cell.font = SUBTOTAL_FONT
    elif kind == "section_total":
        cell.fill = SECTION_TOTAL_FILL
        cell.font = SECTION_TOTAL_FONT
        cell.border = BORDER_TOTAL_TOP


def _write_currency_row(ws, r, col, idr, usd_rate, inr_rate, fill, font):
    for j, val in enumerate((idr, idr / usd_rate if usd_rate else 0,
                             idr / inr_rate if inr_rate else 0)):
        cell = ws.cell(row=r, column=col + j, value=val if val else None)
        cell.number_format = '#,##0;(#,##0);"-"'
        cell.fill = fill
        cell.font = font
        cell.border = BORDER_ALL
        cell.alignment = RIGHT


def write_cf_summary(wb, lines, periods, usd_rate, inr_rate, bb_agg,
                     sheet_name="CF Summary", divisor=1, unit_suffix="",
                     all_periods=None, as_of_label="", current_period=None):
    """Write CF Summary sheet with proper borders and table formatting.

    - If `all_periods` differs from `periods`, adds 'YTD as of <date>' column.
    - If `current_period` provided, adds 'MTD as of <date>' column showing
      current month cumulative only."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    show_ytd_asof = bool(all_periods) and set(all_periods) != set(periods)
    show_mtd = bool(current_period)
    extra_cols = (3 if show_ytd_asof else 0) + (3 if show_mtd else 0)
    last_col = 2 + len(periods) * 3 + 3 + extra_cols

    # Title
    title_cell = ws.cell(row=1, column=2, value="PT Prasad Seeds Indonesia")
    title_cell.font = COMPANY_FONT
    title_cell.alignment = CENTER
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_col)
    ws.row_dimensions[1].height = 24

    # Subtitle / rate info
    sub_cell = ws.cell(row=2, column=2,
            value=(f"Cash Flow Statement   |   USD = Rp {int(usd_rate):,}   |   "
                   f"INR = Rp {int(inr_rate)}   |   Generated: "
                   f"{datetime.now():%d %b %Y %H:%M}"
                   + (unit_suffix or "")))
    sub_cell.font = Font(italic=True, color="666666", size=10)
    sub_cell.alignment = CENTER
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=last_col)
    ws.row_dimensions[2].height = 18

    # Header row 3-4 - "Particular" merged across both rows; month labels
    c = ws.cell(row=3, column=2, value="Particular")
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = CENTER; c.border = BORDER_HEADER
    # Merge Particular cell across rows 3+4 (no empty box below header)
    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
    ws.cell(row=4, column=2).fill = HEADER_FILL
    ws.cell(row=4, column=2).border = BORDER_HEADER

    col = 3
    for period in periods:
        c = ws.cell(row=3, column=col, value=month_label(period))
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = CENTER; c.border = BORDER_HEADER
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 2)
        # Apply border to merged cells
        for k in range(col, col + 3):
            ws.cell(row=3, column=k).border = BORDER_HEADER
        col += 3
    c = ws.cell(row=3, column=col, value=f"YTD {periods[-1].split('-')[0]}")
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = CENTER; c.border = BORDER_HEADER
    ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 2)
    for k in range(col, col + 3):
        ws.cell(row=3, column=k).border = BORDER_HEADER

    # Extra YTD-as-of column (includes current partial month)
    if show_ytd_asof:
        col += 3
        c = ws.cell(row=3, column=col, value=f"YTD as of {as_of_label}")
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = CENTER; c.border = BORDER_HEADER
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 2)
        for k in range(col, col + 3):
            ws.cell(row=3, column=k).border = BORDER_HEADER

    # MTD as-of column (current partial month only)
    if show_mtd:
        col += 3
        c = ws.cell(row=3, column=col, value=f"MTD as of {as_of_label}")
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = CENTER; c.border = BORDER_HEADER
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 2)
        for k in range(col, col + 3):
            ws.cell(row=3, column=k).border = BORDER_HEADER

    ws.row_dimensions[3].height = 22

    # Header row 4 - currency sublabels (per month + YTD + optional YTD/MTD as-of)
    n_blocks = len(periods) + 1 + (1 if show_ytd_asof else 0) + (1 if show_mtd else 0)
    col = 3
    for _ in range(n_blocks):
        for j, cur in enumerate(("IDR", "USD", "INR")):
            c = ws.cell(row=4, column=col + j, value=cur)
            c.fill = HEADER_FILL; c.font = HEADER_FONT
            c.alignment = CENTER; c.border = BORDER_HEADER
        col += 3
    ws.row_dimensions[4].height = 18

    # Data rows
    r = 5
    zebra_toggle = False
    for line in lines:
        kind = line["kind"]
        if kind == "blank":
            # Skip blank rows entirely - no empty boxes in Particular column.
            # Section separation handled visually via colored headers.
            continue
        indent = line.get("indent", 0)
        label_cell = ws.cell(row=r, column=2,
                             value=("    " * indent) + line["label"])
        label_cell.alignment = LEFT
        _apply_style(label_cell, kind)

        ytd = {"IDR": 0.0, "USD": 0.0, "INR": 0.0}
        col = 3
        for period in periods:
            idr = (line["values"].get(period, 0.0) or 0.0) / divisor
            usd = idr / usd_rate if usd_rate else 0
            inr = idr / inr_rate if inr_rate else 0
            ytd["IDR"] += idr; ytd["USD"] += usd; ytd["INR"] += inr
            for j, val in enumerate((idr, usd, inr)):
                cell = ws.cell(row=r, column=col + j,
                               value=val if val else None)
                cell.number_format = '#,##0;(#,##0);"-"'
                cell.alignment = RIGHT
                _apply_style(cell, kind)
            col += 3
        for j, cur in enumerate(("IDR", "USD", "INR")):
            cell = ws.cell(row=r, column=col + j,
                           value=ytd[cur] if ytd[cur] else None)
            cell.number_format = '#,##0;(#,##0);"-"'
            cell.alignment = RIGHT
            _apply_style(cell, kind)

        # YTD as-of column (sum across ALL periods, includes current month)
        if show_ytd_asof:
            col += 3
            ytd_all = {"IDR": 0.0, "USD": 0.0, "INR": 0.0}
            for p in all_periods:
                idr_a = (line["values"].get(p, 0.0) or 0.0) / divisor
                ytd_all["IDR"] += idr_a
                ytd_all["USD"] += idr_a / usd_rate if usd_rate else 0
                ytd_all["INR"] += idr_a / inr_rate if inr_rate else 0
            for j, cur in enumerate(("IDR", "USD", "INR")):
                cell = ws.cell(row=r, column=col + j,
                               value=ytd_all[cur] if ytd_all[cur] else None)
                cell.number_format = '#,##0;(#,##0);"-"'
                cell.alignment = RIGHT
                _apply_style(cell, kind)

        # MTD as-of column (current partial month only)
        if show_mtd:
            col += 3
            idr_m = (line["values"].get(current_period, 0.0) or 0.0) / divisor
            usd_m = idr_m / usd_rate if usd_rate else 0
            inr_m = idr_m / inr_rate if inr_rate else 0
            for j, val in enumerate((idr_m, usd_m, inr_m)):
                cell = ws.cell(row=r, column=col + j, value=val if val else None)
                cell.number_format = '#,##0;(#,##0);"-"'
                cell.alignment = RIGHT
                _apply_style(cell, kind)

        # Zebra striping for leaf rows
        if kind == "leaf" and zebra_toggle:
            for k in range(2, last_col + 1):
                if ws.cell(row=r, column=k).fill.fgColor.rgb in (None, "00000000"):
                    ws.cell(row=r, column=k).fill = ZEBRA_FILL
        if kind == "leaf":
            zebra_toggle = not zebra_toggle
        else:
            zebra_toggle = False
        r += 1

    # Reconciliation block (Beginning + Ending) - placed directly under
    # the last data row, no empty gap above
    incoming_sum = defaultdict(float)
    outflow_sum = defaultdict(float)
    for line in lines:
        if line["label"] == "TOTAL INCOMING":
            for p, v in line["values"].items():
                incoming_sum[p] += v
        if line["label"] == "TOTAL OUTFLOW":
            for p, v in line["values"].items():
                outflow_sum[p] += v

    # BEGINNING row
    cell = ws.cell(row=r, column=2, value="BEGINNING BALANCE")
    cell.fill = SECTION_FILL; cell.font = SECTION_FONT
    cell.alignment = LEFT; cell.border = BORDER_ALL
    col = 3
    for period in periods:
        opening = sum(v for (b, p), v in bb_agg.items() if p == period) / divisor
        _write_currency_row(ws, r, col, opening, usd_rate, inr_rate,
                            SECTION_FILL, SECTION_FONT)
        col += 3
    first_opening = sum(v for (b, p), v in bb_agg.items()
                        if p == periods[0]) / divisor
    _write_currency_row(ws, r, col, first_opening, usd_rate, inr_rate,
                        SECTION_FILL, SECTION_FONT)
    # YTD as-of column for BEGINNING: same first-period opening
    if show_ytd_asof:
        col += 3
        _write_currency_row(ws, r, col, first_opening, usd_rate, inr_rate,
                            SECTION_FILL, SECTION_FONT)
    # MTD as-of BEGINNING: balance at start of current month
    # (= first-period opening + all prior periods net changes)
    if show_mtd:
        col += 3
        first_open_full = sum(v for (b, p), v in bb_agg.items()
                              if p == all_periods[0])
        prior_periods = [p for p in all_periods if p < current_period]
        prior_net = sum(incoming_sum.get(p, 0) + outflow_sum.get(p, 0)
                        for p in prior_periods)
        mtd_open = (first_open_full + prior_net) / divisor
        _write_currency_row(ws, r, col, mtd_open, usd_rate, inr_rate,
                            SECTION_FILL, SECTION_FONT)
    r += 1

    # ENDING row
    cell = ws.cell(row=r, column=2, value="ENDING BALANCE")
    cell.fill = SECTION_TOTAL_FILL; cell.font = SECTION_TOTAL_FONT
    cell.alignment = LEFT; cell.border = BORDER_TOTAL_TOP
    col = 3
    ytd_ending = 0.0
    for period in periods:
        opening = sum(v for (b, p), v in bb_agg.items() if p == period)
        net = incoming_sum.get(period, 0) + outflow_sum.get(period, 0)
        ending = (opening + net) / divisor
        ytd_ending = ending
        for j, val in enumerate((ending, ending / usd_rate if usd_rate else 0,
                                 ending / inr_rate if inr_rate else 0)):
            cc = ws.cell(row=r, column=col + j, value=val if val else None)
            cc.number_format = '#,##0;(#,##0);"-"'
            cc.fill = SECTION_TOTAL_FILL; cc.font = SECTION_TOTAL_FONT
            cc.border = BORDER_TOTAL_TOP; cc.alignment = RIGHT
        col += 3
    for j, val in enumerate((ytd_ending, ytd_ending / usd_rate if usd_rate else 0,
                             ytd_ending / inr_rate if inr_rate else 0)):
        cc = ws.cell(row=r, column=col + j, value=val if val else None)
        cc.number_format = '#,##0;(#,##0);"-"'
        cc.fill = SECTION_TOTAL_FILL; cc.font = SECTION_TOTAL_FONT
        cc.border = BORDER_TOTAL_TOP; cc.alignment = RIGHT

    # YTD as-of ENDING: ending of the latest period in all_periods
    if show_ytd_asof:
        col += 3
        first_opening_full = sum(v for (b, p), v in bb_agg.items() if p == all_periods[0])
        inc_all = sum(incoming_sum.get(p, 0) for p in all_periods)
        out_all = sum(outflow_sum.get(p, 0) for p in all_periods)
        end_all = (first_opening_full + inc_all + out_all) / divisor
        for j, val in enumerate((end_all, end_all / usd_rate if usd_rate else 0,
                                 end_all / inr_rate if inr_rate else 0)):
            cc = ws.cell(row=r, column=col + j, value=val if val else None)
            cc.number_format = '#,##0;(#,##0);"-"'
            cc.fill = SECTION_TOTAL_FILL; cc.font = SECTION_TOTAL_FONT
            cc.border = BORDER_TOTAL_TOP; cc.alignment = RIGHT

    # MTD as-of ENDING: MTD opening + this month's net (= current real-time balance)
    if show_mtd:
        col += 3
        first_open_full = sum(v for (b, p), v in bb_agg.items() if p == all_periods[0])
        prior_periods = [p for p in all_periods if p < current_period]
        prior_net = sum(incoming_sum.get(p, 0) + outflow_sum.get(p, 0)
                        for p in prior_periods)
        cur_net = incoming_sum.get(current_period, 0) + outflow_sum.get(current_period, 0)
        mtd_end = (first_open_full + prior_net + cur_net) / divisor
        for j, val in enumerate((mtd_end, mtd_end / usd_rate if usd_rate else 0,
                                 mtd_end / inr_rate if inr_rate else 0)):
            cc = ws.cell(row=r, column=col + j, value=val if val else None)
            cc.number_format = '#,##0;(#,##0);"-"'
            cc.fill = SECTION_TOTAL_FILL; cc.font = SECTION_TOTAL_FONT
            cc.border = BORDER_TOTAL_TOP; cc.alignment = RIGHT

    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 48
    for cidx in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(cidx)].width = 16
    ws.freeze_panes = "C5"

    # Print setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True


def write_bank_sheets(wb, rows, bb_agg, net_change_agg, periods,
                      usd_rate, inr_rate):
    bank_meta = {}
    for r in rows:
        if r["bank"] not in bank_meta:
            bank_meta[r["bank"]] = {"unit": r["unit"]}

    for period in periods:
        sheet_name = f"Bank - {INDONESIAN_MONTHS[int(period.split('-')[1])]}"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        # Title block
        ws["A1"] = "PT Prasad Seeds Indonesia"
        ws["A1"].font = Font(bold=True, size=14, color="1F3864")
        ws.merge_cells("A1:H1")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 24

        ws["A2"] = "Bank Position Summary"
        ws["A2"].font = Font(bold=True, italic=True, size=11)
        ws.merge_cells("A2:H2")
        ws["A2"].alignment = Alignment(horizontal="center")

        y, m = period.split("-")
        last_day = monthrange(int(y), int(m))[1]
        ws["A3"] = f"As of: {MONTHS_FULL[int(m)]} {last_day}, {y}"
        ws["A3"].font = Font(italic=True, color="666666", size=10)
        ws.merge_cells("A3:H3")
        ws["A3"].alignment = Alignment(horizontal="center")

        # Header row 5
        headers = ["Bank", "Unit", "Currency", "Opening Balance", "Net Change",
                   "Ending Balance (IDR)", "Ending (USD)", "Ending (INR)"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=5, column=i, value=h)
            c.fill = HEADER_FILL; c.font = HEADER_FONT
            c.alignment = CENTER; c.border = BORDER_HEADER
        ws.row_dimensions[5].height = 26

        # Data rows
        r = 6
        total_idr = 0.0
        zebra_toggle = False
        for bank in sorted(bank_meta.keys()):
            opening = bb_agg.get((bank, period), 0.0)
            net_change = net_change_agg.get((bank, period), 0.0)
            ending = opening + net_change
            if not (opening or net_change or ending):
                continue
            row_fill = ZEBRA_FILL if zebra_toggle else None
            for col_idx, val in [(1, bank),
                                  (2, bank_meta[bank]["unit"] or "-"),
                                  (3, "USD-equiv (IDR)" if "USD" in bank else "IDR")]:
                cell = ws.cell(row=r, column=col_idx, value=val)
                cell.border = BORDER_ALL
                cell.alignment = LEFT
                if row_fill: cell.fill = row_fill
            for col_idx, val in [(4, opening), (5, net_change), (6, ending)]:
                cell = ws.cell(row=r, column=col_idx, value=val)
                cell.number_format = '#,##0;(#,##0);"-"'
                cell.border = BORDER_ALL
                cell.alignment = RIGHT
                if row_fill: cell.fill = row_fill
            cell = ws.cell(row=r, column=7, value=ending / usd_rate)
            cell.number_format = '#,##0.00;(#,##0.00);"-"'
            cell.border = BORDER_ALL; cell.alignment = RIGHT
            if row_fill: cell.fill = row_fill
            cell = ws.cell(row=r, column=8, value=ending / inr_rate)
            cell.number_format = '#,##0.00;(#,##0.00);"-"'
            cell.border = BORDER_ALL; cell.alignment = RIGHT
            if row_fill: cell.fill = row_fill
            total_idr += ending
            zebra_toggle = not zebra_toggle
            r += 1

        # Total row
        cell = ws.cell(row=r, column=1, value="TOTAL")
        cell.alignment = LEFT
        for col_idx in range(1, 9):
            c = ws.cell(row=r, column=col_idx)
            c.fill = TOTAL_FILL; c.font = TOTAL_FONT
            c.border = BORDER_TOTAL_TOP
        ws.cell(row=r, column=6, value=total_idr).number_format = '#,##0;(#,##0);"-"'
        ws.cell(row=r, column=6).alignment = RIGHT
        ws.cell(row=r, column=7, value=total_idr / usd_rate).number_format = '#,##0.00;(#,##0.00);"-"'
        ws.cell(row=r, column=7).alignment = RIGHT
        ws.cell(row=r, column=8, value=total_idr / inr_rate).number_format = '#,##0.00;(#,##0.00);"-"'
        ws.cell(row=r, column=8).alignment = RIGHT
        # Re-apply borders + fonts after setting values
        for col_idx in range(1, 9):
            c = ws.cell(row=r, column=col_idx)
            c.fill = TOTAL_FILL; c.font = TOTAL_FONT
            c.border = BORDER_TOTAL_TOP

        # Column widths
        for cidx, w in enumerate([20, 12, 18, 22, 22, 24, 18, 18], start=1):
            ws.column_dimensions[get_column_letter(cidx)].width = w

        ws.freeze_panes = "A6"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
