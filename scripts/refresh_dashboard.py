"""
PSI Cash Flow Dashboard - Auto Refresh
=======================================
PT Prasad Seeds Indonesia

Reads "PSI Cash Monitoring Master.xlsx" and:
  1. Regenerates the "CF Summary" sheet (replaces manual pivot).
  2. Regenerates the "Bank - <Month>" sheets.
  3. Writes public/data.json + public/data.js (optionally encrypted).
"""
from __future__ import annotations

import json
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from lib_pivot import (
    read_all_banks, detect_currency_rates, all_periods,
    aggregate_by_path, aggregate_beginning_balance,
    aggregate_net_change_per_bank, build_cf_structure,
    aggregate_beginning_balance_usd, aggregate_net_change_per_bank_usd,
)
from lib_excel import write_cf_summary, write_bank_sheets
from lib_dashboard import build_dashboard_data

ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSX = ROOT / "PSI Cash Monitoring Master.xlsx"
PUBLIC_DIR = ROOT / "public"
DATA_JSON = PUBLIC_DIR / "data.json"
DATA_JS = PUBLIC_DIR / "data.js"
PASSWORD_FILE = ROOT / "dashboard-password.txt"


def is_excel_open(path: Path) -> bool:
    return (path.parent / f"~${path.name}").exists()


def main():
    src = Path(os.environ.get("PSI_SOURCE_XLSX", str(SOURCE_XLSX)))
    if not src.exists():
        print(f"ERROR: Source file not found: {src}")
        sys.exit(1)

    if is_excel_open(src):
        print("=" * 70)
        print("WARNING: It looks like the Excel file is currently OPEN in Excel")
        print(f"        ({src.name}).")
        print("        Please close Excel and run this script again.")
        print("=" * 70)
        sys.exit(2)

    print(f"Reading {src.name} ...")
    wb = openpyxl.load_workbook(src, data_only=True)
    rows = read_all_banks(wb)

    # ONE-TIME CLEANUP: remove stale auto-generated sheets from source.
    # Previous versions of this script wrote CF Summary + Bank-Month sheets
    # directly into the source workbook. Now they belong in the separate
    # PSI Cash Flow Report.xlsx, so we clean the source to keep only the
    # treasury's working sheets (All Banks etc).
    stale = []
    for sn in list(wb.sheetnames):
        if sn in ("CF Summary", "CF Summary (IDR Mio)"):
            stale.append(sn)
        elif sn.startswith("Bank - "):
            stale.append(sn)
    if stale:
        print(f"  Found {len(stale)} stale auto-gen sheets in source. Cleaning ...")
        wb_clean = openpyxl.load_workbook(src)
        for sn in stale:
            if sn in wb_clean.sheetnames:
                del wb_clean[sn]
                print(f"    Removed: {sn}")
        # Backup first
        backup_dir = ROOT / "backups"
        backup_dir.mkdir(exist_ok=True)
        bk = backup_dir / f"PSI Cash Monitoring Master.cleanup.{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        shutil.copy2(src, bk)
        wb_clean.save(src)
        print(f"  Source cleaned. Backup: {bk.name}")
        # Reload source after cleanup so subsequent reads are clean
        wb = openpyxl.load_workbook(src, data_only=True)
    print(f"  {len(rows):,} transactions loaded")

    usd_rate, inr_rate = detect_currency_rates(wb)
    print(f"  Rates: USD = Rp {usd_rate:,.0f}  |  INR = Rp {inr_rate:,.0f}")

    periods = all_periods(rows)
    print(f"  Periods found: {', '.join(periods)}")

    agg = aggregate_by_path(rows)
    bb_agg = aggregate_beginning_balance(rows)
    net_change_agg = aggregate_net_change_per_bank(rows)
    bb_agg_usd = aggregate_beginning_balance_usd(rows)
    net_change_agg_usd = aggregate_net_change_per_bank_usd(rows)

    # CF Summary in Excel only includes COMPLETED months (current month
    # excluded so monthly reports aren't polluted by partial-month data).
    today_period = f"{datetime.now().year:04d}-{datetime.now().month:02d}"
    completed_periods = [p for p in periods if p < today_period]
    print(f"  Today = {today_period}. Completed months for Excel report: "
          f"{', '.join(completed_periods) if completed_periods else '(none)'}")

    # Build CF Summary structure (completed months only for Excel report)
    print("Building report structures ...")
    completed_lines = build_cf_structure(agg, completed_periods, rows=rows) if completed_periods else []

    # Dashboard always uses ALL periods (Daily View needs current month).
    lines = build_cf_structure(agg, periods, rows=rows)

    # ====================================================================
    # Generate SEPARATE report file (PSI Cash Flow Report.xlsx) in public/
    # Source PSI Cash Monitoring Master.xlsx is NOT modified anymore.
    # ====================================================================
    import openpyxl as _ox
    wb_report = _ox.Workbook()
    if "Sheet" in wb_report.sheetnames:
        wb_report.remove(wb_report["Sheet"])

    if completed_periods:
        today_str = datetime.now().strftime("%d %b %Y")
        print("Writing CF Summary to report (completed months + YTD as-of) ...")
        # Use `lines` (built from all periods) so YTD-as-of includes partial current month
        write_cf_summary(wb_report, lines, completed_periods,
                         usd_rate, inr_rate, bb_agg,
                         all_periods=periods, as_of_label=today_str)
        print("Writing CF Summary (IDR Mio) to report ...")
        write_cf_summary(wb_report, lines, completed_periods,
                         usd_rate, inr_rate, bb_agg,
                         sheet_name="CF Summary (IDR Mio)", divisor=1_000_000,
                         unit_suffix=" (in IDR Million)",
                         all_periods=periods, as_of_label=today_str)
    else:
        print("[INFO] No completed months yet - CF Summary sheets skipped.")

    print("Writing Bank - <Month> sheets to report (all periods) ...")
    write_bank_sheets(wb_report, rows, bb_agg, net_change_agg, periods,
                      usd_rate, inr_rate)

    PUBLIC_DIR.mkdir(exist_ok=True)
    REPORT_PATH = PUBLIC_DIR / "PSI Cash Flow Report.xlsx"
    wb_report.save(REPORT_PATH)
    print(f"  Saved report: {REPORT_PATH.relative_to(ROOT)}")
    print(f"  Source NOT modified: {src.name}")

    print("Building dashboard data ...")
    data = build_dashboard_data(rows, agg, bb_agg, net_change_agg,
                                lines, periods, usd_rate, inr_rate,
                                bb_agg_usd=bb_agg_usd,
                                net_change_agg_usd=net_change_agg_usd)
    PUBLIC_DIR.mkdir(exist_ok=True)
    json_text = json.dumps(data, indent=2, default=str)

    password = None
    if PASSWORD_FILE.exists():
        pw = PASSWORD_FILE.read_text(encoding="utf-8").strip()
        if pw:
            password = pw

    if password:
        from lib_crypto import encrypt_payload
        print(f"  Encrypting payload with {PASSWORD_FILE.name} ...")
        enc = encrypt_payload(json_text, password)
        enc_text = json.dumps(enc, indent=2)
        DATA_JSON.write_text(enc_text, encoding="utf-8")
        DATA_JS.write_text(
            "// Auto-generated. Encrypted; password required to decrypt.\n"
            "window.PSI_DATA_ENCRYPTED = " + enc_text + ";\n",
            encoding="utf-8")
        print(f"  Saved (encrypted): {DATA_JSON.relative_to(ROOT)}")
        print(f"  Saved (encrypted): {DATA_JS.relative_to(ROOT)}")
    else:
        DATA_JSON.write_text(json_text, encoding="utf-8")
        DATA_JS.write_text(
            "// Auto-generated. Plain (no password set in "
            "dashboard-password.txt).\n"
            "window.PSI_DATA = " + json_text + ";\n",
            encoding="utf-8")
        print(f"  Saved (plain): {DATA_JSON.relative_to(ROOT)}")
        print(f"  Saved (plain): {DATA_JS.relative_to(ROOT)}")
        print("  NOTE: No password file found. Anyone with URL can view.")
        print(f"        To enable password: create '{PASSWORD_FILE.name}'")
        print("        in this folder containing your password (1 line).")

    print("\nRefresh OK. Open public/index.html to view dashboard.")


if __name__ == "__main__":
    main()
